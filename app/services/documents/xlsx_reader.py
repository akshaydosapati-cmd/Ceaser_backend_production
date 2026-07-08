from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.services.documents.schemas import ExtractedDocument


class XLSXReader:
    def read(self, path: Path) -> ExtractedDocument:
        workbook = load_workbook(str(path), read_only=True, data_only=True)
        chunks = []
        for sheet in workbook.worksheets:
            rows = []
            for row in sheet.iter_rows(values_only=True):
                values = [str(value) for value in row if value is not None]
                if values:
                    rows.append(" | ".join(values))
            if rows:
                chunks.append(f"Sheet: {sheet.title}\n" + "\n".join(rows[:200]))
        return ExtractedDocument(title=path.stem, pages=len(workbook.worksheets), content="\n\n".join(chunks), metadata={"reader": "openpyxl"})
